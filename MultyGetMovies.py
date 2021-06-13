import time
import requests
import lxml.html
import re
import os
import xlrd
from multiprocessing.dummy import Pool
import openpyxl


# html = requests.get('https://yts.mx/torrent/download/363BC6C534B1430C6758318D196CCD61DB61B647').content
# print(html)
# with open('the maritx.torrent','wb') as f:
#     f.write(html)
# movie_name = ['The Matrix (1999)', 'The Matrix Revolutions (2003)', 'The Matrix Reloaded (2003)']

'''
    因为我的Excel中电影名都是从豆瓣直接爬取的，所以要满足中文名+英文名+年份的要求。有自己特殊的格式的话可以在get_colum()方法自己改
    如：黑客帝国的根源 The roots of the Matrix (2005)  可以
       The Matrix           不可以
       黑客帝国               不可以
       黑客帝国的根源 The roots of the Matrix 不可以
'''
movie_name = []
def get_colum(): #从Excel表格中获取要批量下载的电影的名字
    book = xlrd.open_workbook('电影.xlsx')
    sheet = book.sheet_by_index(0)#Excel表格的第一个Sheet
    values = sheet.col_values(colx=0, start_rowx=1)#从第1列，第2行开始读取。可按需更改
    print(values)
    # print(len(values))
    for i in range(len(values)):
        name = re.findall('[\u4e00-\u9fa5] (.*\(\d.*\))', values[i]) # 正则表达式去除汉字
        if name != []:
            dict_name = {values[i]:name[0]}
            movie_name.append(dict_name)
        else:
            continue
    return movie_name

# def get_values():
#     book = xlrd.open_workbook('搜索不到的高分美国电影.xlsx')
#     sheet = book.sheet_by_index(0)
#     values = sheet.col_values(colx=0, start_rowx=1)
#     return values
#
# values = get_values()
# print(values)
down_list = []
no_list = []
no1080p_list = []
def get_movies(movie_name):
    for key,name in movie_name.items():
        # 获取影片列表
        url = 'https://yts.mx/browse-movies/{}/all/all/0/latest/0/all'.format(name)
        html = requests.get(url).content.decode()
        selector = lxml.html.fromstring(html)
        number = selector.xpath('/ html / body / div[4] / div[4] / div / h2 / b /text()')  # 获取页面搜索结果
        if number != []:  # 首先保证不是一个空列表
            number = int(number[0])  # 转换成整型
            if number == 1:  # 等于1说明有结果
                title_url = selector.xpath('/html/body/div[4]/div[4]/div/section/div/div/a/@href')  # 获取影片的详情页
                if title_url != []:
                    href = title_url[0]  # 获取影片详情页链接
                    movie_content = requests.get(href).content.decode()  # 获取影片详情页的源代码
                    findall = re.findall('<p class="hidden-md hidden-lg">(.*?)</p>', movie_content, re.S)  # 获取所有下载链接
                    if findall != []:
                        re_findall = re.findall('<a.*?</a>', findall[0])  # 将所有下载链接组成一个列表
                    else:
                        continue
                    for x in range(len(re_findall)):  # 循环下载链接，找出1080p的下载链接
                        type = re.findall('</span>(.*?)</a>', re_findall[x])  # 链接的文本
                        torror = re.findall('<a href="(.*?)"', re_findall[x])  # 下载链接
                        if type[0] == '1080p.BluRay':  # 如果是1080p就下载
                            torror_url = torror[0]
                            html = requests.get(torror_url).content
                            os.makedirs('电影', exist_ok=True)#目录名
                            file_path = os.path.join('电影', name + '.torrent')#文件名
                            try:
                                with open(file_path, 'wb') as f:
                                    f.write(html)
                            except:
                                continue
                            down_list.append(key)
                            w_down_list(down_list)
                                # print(file_path+'1080p.BluRay'+"写入成功")
                            break
                        else:
                            if type[0] == '1080p.WEB':
                                torror_url = torror[0]
                                html = requests.get(torror_url).content
                                os.makedirs('电影', exist_ok=True)#目录名
                                file_path = os.path.join('电影', name + '.torrent')#文件名
                                try:
                                    with open(file_path, 'wb') as f:
                                        f.write(html)
                                except:
                                    continue
                                down_list.append(key)
                                break
                                    # print(file_path+'1080p.WEB'+"写入成功")
                            else:
                                # if type[0] == '720p.BluRay'or type[0] == '720p.Web'and x==len(re_findall)-1:
                                #     no1080p_list.append(key)
                                # else:
                                 continue


                else:
                    print("无法获取{}的链接".format(name))

            else:
                no_list.append(key)
                print('{}：本页面无结果'.format(name))
        else:  # 等于0说明没有此影片的数据
            print("无法获取数据")


# def w_tor(torror_url):  # 保存下载文件
#     html = requests.get(torror_url).content
#     with open('the maritx.torrent', 'wb') as f:
#         f.write(html)


'''
    这三个方法分别生成了三个表格：
    1.成功下载的电影
    2.搜索不到的电影
    3.没有1080p资源的电影
'''
def w_down_list(a):
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = '高分电影'
    sheet['A1'] = '电影名：'
    row = 2
    for i in range(len(a)):
        sheet.cell(row,1,a[i])
        row=row+1
    book.save('电影_Down.xlsx')

def w_no_list(a):
    book_1 = openpyxl.Workbook()
    sheet = book_1.active
    sheet.title = '爱情'
    sheet['A1'] = '电影名：'
    row = 2
    for i in range(len(a)):
        sheet.cell(row,1,a[i])
        row=row+1
    book_1.save('搜索不到的电影.xlsx')

def w_no1080p_list(a):
    book_2 = openpyxl.Workbook()
    sheet = book_2.active
    sheet.title = '爱情'
    sheet['A1'] = '电影名：'
    row = 2
    for i in range(len(a)):
        sheet.cell(row,1,a[i])
        row=row+1
    book_2.save('没有1080p的电影.xlsx')

if __name__ == '__main__':
    # start_time = time.time()#开始时间
    movie_name = get_colum()# 获取要下载的电影的名字
    # print(movie_name)
    pool = Pool(200)# 开两百个线程批量下载
    pool.map(get_movies,movie_name)
    w_down_list(down_list)#成功下载的电影的表格
    w_no_list(no_list)#搜索不到的电影的表格
    w_no1080p_list(no1080p_list)#没有1080p的电影的表格
    # end_time = time.time()
    # print("耗时:{}".format(end_time-start_time))
